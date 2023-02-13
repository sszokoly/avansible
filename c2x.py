import csv
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

FONT_BOLD = Font(name='Calibri', size=11, bold=True)
FG_GRAY = PatternFill(patternType='solid', fgColor='E1E1E1')
FG_ORNG = PatternFill(patternType='solid', fgColor='FF8E4B')
SIDE_THIN = Side(border_style='thin', color='000000')
SIDE_DOUBLE = Side(border_style='double', color='000000')
BORDER_THIN = Border(top=SIDE_THIN, bottom=SIDE_THIN, right=SIDE_THIN, left=SIDE_THIN)
BORDER_DBL_BOTTOM = Border(top=SIDE_THIN, bottom=SIDE_DOUBLE, right=SIDE_THIN, left=SIDE_THIN)
ALIGN_HOR_CENTER = Alignment(horizontal='center')


def load_csv(csvfile, ws, delimiter=','):
    with open(csvfile, encoding='utf-8') as fd:
        reader = csv.reader(fd, delimiter=delimiter)
        for row in reader:
            ws.append(row)
    return ws

def adjust_column_widths(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0),
                     len(str(cell.value))))
    for column_letter, width in dims.items():
        ws.column_dimensions[column_letter].width = width + 4
    return ws

def style_header(ws):
    for col in ws.iter_cols(max_row=1):
        col[0].font = Font(name='Calibri', size=11, color='FFFFFF')
        col[0].fill = PatternFill(patternType='solid', fgColor='1C1C1C')
        col[0].alignment = Alignment(horizontal='center')
    return ws

def style_products(ws):
    for row in list(ws.rows)[1:]:
        if row[0].value.isupper():
            row[0].fill = PatternFill(patternType='solid', fgColor='FF8E4B')
            row[0].border = BORDER_THIN
    return ws

def load_ws(wb, inxlsx, sheetname):
    if sheetname:
        if sheetname not in wb.sheetnames:
            if inxlsx:
                wb.create_sheet(sheetname)
                ws = wb[sheetname]
            else:
                ws = wb.active
                ws.title = sheetname
        else:
            index = wb.sheetnames.index(sheetname)
            wb.remove(wb[sheetname])
            wb.create_sheet(sheetname, index=index)
            ws = wb[sheetname]
    else:
        ws = wb.active
    return ws

def strip_extension(filename):
    path, filename_with_ext = os.path.split(filename)
    root, _ = os.path.splitext(filename_with_ext)
    return os.path.join(path, root)

def horizontal_align_columns(ws, columns=None, alignment='center'):
    al = Alignment(horizontal=alignment)
    if not columns:
        columns = range(1, ws.max_column + 1)
    for column in columns:
        for cell in ws[get_column_letter(column)]:
            cell.alignment = al
    return ws

def horizontal_align_rows(ws, rows=None, alignment='center'):
    al = Alignment(horizontal=alignment)
    if not rows:
        rows = range(1, ws.max_rows + 1)
    for row in rows:
        for cell in ws[row]:
            cell.alignment = al
    return ws

def main(csvfile, inxlsx=None, sheetname=None):
    wb = load_workbook(inxlsx) if inxlsx else Workbook()
    ws = load_ws(wb, inxlsx, sheetname)
    ws = load_csv(csvfile, ws)
    ws = adjust_column_widths(ws)
    ws = horizontal_align_columns(ws, columns=[3, 4], alignment='center')
    ws = horizontal_align_columns(ws, columns=[1], alignment='left')
    ws = horizontal_align_columns(ws, columns=[2, 5, 6, 7, 8], alignment='right')
    ws = horizontal_align_rows(ws, rows=[1], alignment='center')
    ws = style_header(ws)
    ws = style_products(ws)
    outxlsx = strip_extension(csvfile) + '.xlsx' if not inxlsx else inxlsx
    wb.save(outxlsx)

if __name__ == '__main__':
    csvfile = 'swversions.csv'
    main(csvfile, sheetname='swversions')
